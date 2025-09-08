from flask import Flask, request, jsonify, send_from_directory
import requests
from openpyxl import load_workbook
from io import BytesIO
import base64
import os
from cryptography.fernet import Fernet

app = Flask(__name__)

# ===== CONFIG =====
ENCRYPTED_TOKEN = b"gAAAAABnIU7Oj2-0mFVw9QUSCeV--QFpnYBeaH7pN9l94KRVXh1rXEl6l9n9GClp4l7RoDzKnDxX_0r1x4Q4oqmUraH9UOYzm2VG_5zzjfxIf2sVzF8mY4wGpbwR9JQMyjEan_jOiMxX"
SECRET_KEY = os.getenv("7DyACMXHCee3H4UgL_UxA0b80tUibcws6sAVs3VGjX8=")

if SECRET_KEY:
    fernet = Fernet(SECRET_KEY.encode())
    GITHUB_TOKEN = fernet.decrypt(ENCRYPTED_TOKEN).decode()
else:
    GITHUB_TOKEN = None

REPO = "supun123456789/job-tracker"  # Replace with your repo
BRANCH = "main"
FILE_PATH = "jobs.xlsx"
GITHUB_API_URL = f"https://api.github.com/repos/{REPO}/contents/{FILE_PATH}"
# ==================

# Serve frontend
@app.route("/")
def home():
    return send_from_directory(".", "index.html")

# Submit job route
@app.route('/submit_job', methods=['POST'])
def submit_job():
    if not GITHUB_TOKEN:
        return jsonify({"status":"error","message":"GitHub token missing or not decrypted!"}),500
    try:
        job_data = request.json
        file_stream, sha = get_excel_file()
        updated_file = update_excel(job_data, file_stream)
        push_to_github(updated_file, sha)
        return jsonify({"status":"success"})
    except Exception as e:
        return jsonify({"status":"error","message":str(e)}),500

# Global error handler
@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({"status":"error","message":str(e)}),500

# Excel helper functions
def get_excel_file():
    headers = {"Authorization": f"token {GITHUB_TOKEN}"}
    r = requests.get(GITHUB_API_URL + f"?ref={BRANCH}", headers=headers)
    r.raise_for_status()
    data = r.json()
    sha = data['sha']
    content = base64.b64decode(data['content'])
    return BytesIO(content), sha

def update_excel(job_data, file_stream):
    wb = load_workbook(file_stream)
    ws = wb.active

    job_numbers = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row+1)]
    if job_data['job_number'] in job_numbers:
        row_idx = job_numbers.index(job_data['job_number']) + 2
    else:
        row_idx = ws.max_row + 1

    ws.cell(row=row_idx, column=1, value=job_data['job_number'])
    ws.cell(row=row_idx, column=2, value=job_data['customer_name'])
    ws.cell(row=row_idx, column=3, value=job_data['job_state'])
    ws.cell(row=row_idx, column=4, value=job_data['job_in_time'])
    ws.cell(row=row_idx, column=5, value=job_data['job_out_time'])
    ws.cell(row=row_idx, column=6, value=job_data['remark'])

    out_stream = BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)
    return out_stream

def push_to_github(file_stream, sha):
    content = base64.b64encode(file_stream.read()).decode('utf-8')
    headers = {"Authorization": f"token {GITHUB_TOKEN}"}
    data = {
        "message": "Update jobs.xlsx",
        "content": content,
        "branch": BRANCH,
        "sha": sha
    }
    r = requests.put(GITHUB_API_URL, headers=headers, json=data)
    r.raise_for_status()
    return r.json()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT",5000)))

